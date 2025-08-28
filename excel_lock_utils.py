"""
Excel file locking utilities for safe concurrent access
Uses file-based locking to prevent concurrent writes
"""

import asyncio
import fcntl
import time
import os
from contextlib import contextmanager
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

from config import EXCEL_FILE_PATH, UAE_TZ
from logger import logger


class ExcelLockManager:
    """Manages file locks for Excel operations"""
    
    def __init__(self, lock_timeout: int = 30):
        self.lock_timeout = lock_timeout
        self.lock_file = f"{EXCEL_FILE_PATH}.lock"
        self.retry_delay = 0.1  # 100ms between retries
        
    @contextmanager
    def acquire_lock(self, operation: str = "write"):
        """Acquire exclusive lock for Excel file operations"""
        lock_fd = None
        start_time = time.time()
        
        try:
            # Create lock file if it doesn't exist
            lock_fd = os.open(self.lock_file, os.O_CREAT | os.O_WRONLY)
            
            # Try to acquire exclusive lock
            while True:
                try:
                    fcntl.flock(lock_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
                    logger.debug(f"Acquired Excel lock for {operation}")
                    break
                except IOError:
                    # Lock is held by another process
                    if time.time() - start_time > self.lock_timeout:
                        raise TimeoutError(f"Could not acquire Excel lock after {self.lock_timeout} seconds")
                    time.sleep(self.retry_delay)
            
            # Write lock info
            lock_info = f"{os.getpid()}:{operation}:{datetime.now(UAE_TZ).isoformat()}\n"
            os.write(lock_fd, lock_info.encode())
            os.fsync(lock_fd)
            
            yield
            
        finally:
            # Release lock
            if lock_fd is not None:
                try:
                    fcntl.flock(lock_fd, fcntl.LOCK_UN)
                    os.close(lock_fd)
                    logger.debug(f"Released Excel lock for {operation}")
                except Exception as e:
                    logger.warning(f"Error releasing lock: {e}")
    
    def is_locked(self) -> bool:
        """Check if Excel file is currently locked"""
        if not os.path.exists(self.lock_file):
            return False
            
        try:
            lock_fd = os.open(self.lock_file, os.O_RDONLY)
            try:
                fcntl.flock(lock_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
                fcntl.flock(lock_fd, fcntl.LOCK_UN)
                os.close(lock_fd)
                return False
            except IOError:
                os.close(lock_fd)
                return True
        except Exception:
            return False
    
    def get_lock_info(self) -> Optional[str]:
        """Get information about current lock holder"""
        if not os.path.exists(self.lock_file):
            return None
            
        try:
            with open(self.lock_file, 'r') as f:
                return f.read().strip()
        except Exception:
            return None


# Global instance
excel_lock = ExcelLockManager()


# Enhanced Excel operations with locking
async def safe_read_excel() -> pd.DataFrame:
    """Read Excel file with shared lock"""
    is_render = os.environ.get("RENDER") == "true"
    
    if is_render:
        # On Render, read without locking
        return await asyncio.to_thread(pd.read_excel, EXCEL_FILE_PATH)
    else:
        # Local environment - use locking
        with excel_lock.acquire_lock("read"):
            return await asyncio.to_thread(pd.read_excel, EXCEL_FILE_PATH)


async def safe_write_excel(df: pd.DataFrame) -> bool:
    """Write to Excel file with exclusive lock"""
    # Check if we're on Render - use simpler approach without fcntl
    is_render = os.environ.get("RENDER") == "true"
    
    if is_render:
        # On Render, use a simpler approach without file locking
        try:
            # Write to temporary file first
            temp_path = f"{EXCEL_FILE_PATH}.tmp"
            await asyncio.to_thread(df.to_excel, temp_path, index=False)
            
            # Format headers
            wb = await asyncio.to_thread(load_workbook, temp_path)
            ws = wb.active
            from openpyxl.styles import Font
            for cell in ws[1]:
                cell.font = Font(bold=True)
            await asyncio.to_thread(wb.save, temp_path)
            wb.close()
            
            # Atomic rename
            os.rename(temp_path, EXCEL_FILE_PATH)
            
            logger.debug("Excel write completed on Render without locking")
            return True
            
        except Exception as e:
            logger.error(f"Error writing Excel on Render: {e}")
            logger.error(f"Excel path: {EXCEL_FILE_PATH}")
            logger.error(f"Temp path: {temp_path}")
            logger.error(f"Current working dir: {os.getcwd()}")
            logger.error(f"Excel file exists: {os.path.exists(EXCEL_FILE_PATH)}")
            logger.error(f"Excel file permissions: {oct(os.stat(EXCEL_FILE_PATH).st_mode) if os.path.exists(EXCEL_FILE_PATH) else 'N/A'}")
            logger.error(f"Directory permissions: {oct(os.stat(os.path.dirname(EXCEL_FILE_PATH)).st_mode)}")
            # Clean up temp file if it exists
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
            return False
    else:
        # Local environment - use full locking
        try:
            with excel_lock.acquire_lock("write"):
                # Write to temporary file first
                temp_path = f"{EXCEL_FILE_PATH}.tmp"
                df.to_excel(temp_path, index=False)
                
                # Format headers
                wb = load_workbook(temp_path)
                ws = wb.active
                from openpyxl.styles import Font
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                wb.save(temp_path)
                
                # Atomic rename
                os.rename(temp_path, EXCEL_FILE_PATH)
                
                return True
                
        except TimeoutError:
            logger.error("Could not acquire lock for Excel write - file may be in use")
            return False
        except Exception as e:
            logger.error(f"Error writing Excel with lock: {e}")
            return False


async def safe_append_row(row_data: list) -> bool:
    """Append a row to Excel file with locking"""
    try:
        with excel_lock.acquire_lock("append"):
            # Read current data
            if os.path.exists(EXCEL_FILE_PATH):
                df = pd.read_excel(EXCEL_FILE_PATH)
            else:
                # Create new DataFrame with headers
                headers = ["Task #", "Timestamp", "Brand", "Campaign Start Date", 
                          "Campaign End Date", "Reference Number", "Location", 
                          "Sales Person", "Submitted By", "Status", "Filming Date", 
                          "Videographer", "Video Filename", "Current Version", 
                          "Version History", "Pending Timestamps", "Submitted Timestamps", 
                          "Returned Timestamps", "Rejected Timestamps", "Accepted Timestamps"]
                df = pd.DataFrame(columns=headers)
            
            # Append new row
            df.loc[len(df)] = row_data
            
            # Write back
            return await safe_write_excel(df)
            
    except Exception as e:
        logger.error(f"Error appending row with lock: {e}")
        return False


# Notification system for lock conflicts
class LockConflictNotifier:
    """Notifies users when Excel operations are blocked"""
    
    def __init__(self):
        self.pending_operations = {}
        
    async def notify_lock_conflict(self, user_id: str, operation: str):
        """Notify user that their operation is waiting for lock"""
        lock_info = excel_lock.get_lock_info()
        
        message = f"⏳ Your {operation} is waiting for Excel access..."
        if lock_info:
            message += f"\nCurrently locked by: {lock_info.split(':')[0]}"
        
        # Send Slack notification
        from clients import slack_client
        try:
            await slack_client.chat_postEphemeral(
                channel=user_id,
                user=user_id,
                text=message
            )
        except Exception as e:
            logger.warning(f"Could not notify user about lock: {e}")
    
    async def notify_operation_complete(self, user_id: str, operation: str, success: bool):
        """Notify user when their operation completes"""
        if success:
            message = f"✅ Your {operation} completed successfully!"
        else:
            message = f"❌ Your {operation} failed - please try again"
        
        from clients import slack_client
        try:
            await slack_client.chat_postEphemeral(
                channel=user_id,
                user=user_id,
                text=message
            )
        except Exception as e:
            logger.warning(f"Could not notify user about completion: {e}")


lock_notifier = LockConflictNotifier()