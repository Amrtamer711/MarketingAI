import asyncio
import os
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from config import EXCEL_FILE_PATH, UAE_TZ, HISTORY_DB_PATH
from logger import logger
from trello_utils import get_trello_card_by_task_number, update_trello_card, get_trello_lists
from utils import calculate_filming_date

# ========== ASYNC EXCEL MANAGEMENT ==========
async def initialize_excel():
    """Create Excel file with headers if it doesn't exist"""
    if not Path(EXCEL_FILE_PATH).exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Design requests"
        
        headers = ["Task #", "Timestamp", "Brand", "Campaign Start Date", "Campaign End Date", "Reference Number", 
                  "Location", "Sales Person", "Submitted By", "Status", "Filming Date", "Videographer", "Video Filename",
                  "Current Version", "Version History",
                  "Pending Timestamps", "Submitted Timestamps", "Returned Timestamps", "Rejected Timestamps", "Accepted Timestamps"]
        ws.append(headers)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Save asynchronously
        await asyncio.to_thread(wb.save, EXCEL_FILE_PATH)

async def save_to_excel(data: Dict[str, Any]) -> Dict[str, Any]:
    """Save parsed campaign data to Excel file
    Returns: {"success": bool, "task_number": int or None}
    """
    try:
        # Create file if it doesn't exist
        if not Path(EXCEL_FILE_PATH).exists():
            await initialize_excel()
        
        # Load workbook in thread pool
        wb = await asyncio.to_thread(load_workbook, EXCEL_FILE_PATH)
        ws = wb.active
        
        # Get next task number
        task_number = await get_next_task_number()
        
        # Calculate filming date with new rules
        filming_date = calculate_filming_date(
            data.get("start_date", ""),
            data.get("end_date", "")
        )
        
        # Format dates to DD-MM-YYYY
        start_date = data.get("start_date", "")
        end_date = data.get("end_date", "")
        
        # Convert dates if they're in YYYY-MM-DD format
        if start_date and len(start_date) == 10 and start_date[4] == '-':
            try:
                date_obj = datetime.strptime(start_date, "%Y-%m-%d")
                start_date = date_obj.strftime("%d-%m-%Y")
            except:
                pass
                
        if end_date and len(end_date) == 10 and end_date[4] == '-':
            try:
                date_obj = datetime.strptime(end_date, "%Y-%m-%d")
                end_date = date_obj.strftime("%d-%m-%Y")
            except:
                pass
        
        # Ensure new columns exist in case of older files
        existing_headers = [c.value for c in ws[1]]
        new_headers = ["Current Version", "Version History"]
        if any(h not in existing_headers for h in new_headers):
            for h in new_headers:
                if h not in existing_headers:
                    ws.cell(row=1, column=len(existing_headers) + 1, value=h)
                    existing_headers.append(h)
            # Also ensure timestamp columns exist
            for h in ["Pending Timestamps", "Submitted Timestamps", "Returned Timestamps", "Rejected Timestamps", "Accepted Timestamps"]:
                if h not in existing_headers:
                    ws.cell(row=1, column=len(existing_headers) + 1, value=h)
                    existing_headers.append(h)
        
        # Replace underscores with dashes in all fields
        row_data = [
            task_number,
            datetime.now(UAE_TZ).strftime("%d-%m-%Y %H:%M:%S"),
            data.get("brand", "").replace("_", "-"),
            start_date,
            end_date,
            data.get("reference_number", "").replace("_", "-"),
            data.get("location", "").replace("_", "-"),
            data.get("sales_person", "").replace("_", "-"),
            data.get("submitted_by", "").replace("_", "-"),
            "Not assigned yet",  # Default status
            filming_date,
            "",  # Videographer - will be filled during assignment
            "",  # Video Filename - will be filled during upload/assignment
            "",  # Current Version
            "[]",  # Version History (JSON array)
            "",  # Pending Timestamps
            "",  # Submitted Timestamps
            "",  # Returned Timestamps
            "",  # Rejected Timestamps
            ""   # Accepted Timestamps
        ]
        
        ws.append(row_data)
        
        # Save asynchronously
        await asyncio.to_thread(wb.save, EXCEL_FILE_PATH)
        
        return {"success": True, "task_number": task_number}
    except Exception as e:
        logger.error(f"Error saving to Excel: {e}")
        return {"success": False, "task_number": None}

async def read_excel_async() -> pd.DataFrame:
    """Read data from Excel file"""
    if not Path(EXCEL_FILE_PATH).exists():
        await initialize_excel()
    return await asyncio.to_thread(pd.read_excel, EXCEL_FILE_PATH)

async def export_current_data(include_history: bool = True, format: str = "summary") -> str:
    """Export all current data from Excel and optionally from history database"""
    try:
        import sqlite3
        from config import HISTORY_DB_PATH
        
        # Get Excel data (live tasks)
        df = await read_excel_async()
        
        response = "📊 **Current System Data Export**\n\n"
        response += "=" * 50 + "\n\n"
        
        # 1. LIVE TASKS FROM EXCEL
        response += f"**📋 LIVE TASKS (Excel) - {len(df)} tasks**\n\n"
        
        if len(df) > 0:
            # Sort by task number descending
            df_sorted = df.sort_values('Task #', ascending=False)
            
            for _, row in df_sorted.iterrows():
                task_num = row.get('Task #', 'N/A')
                status = row.get('Status', 'N/A')
                
                if format == "summary":
                    # Summary format - key fields only
                    response += f"**Task #{task_num}** - {status}\n"
                    response += f"• Brand: {row.get('Brand', 'N/A')}\n"
                    response += f"• Campaign: {row.get('Campaign Start Date', 'N/A')} to {row.get('Campaign End Date', 'N/A')}\n"
                    response += f"• Reference: {row.get('Reference Number', 'N/A')}\n"
                    response += f"• Location: {row.get('Location', 'N/A')}\n"
                    response += f"• Sales Person: {row.get('Sales Person', 'N/A')}\n"
                    if pd.notna(row.get('Videographer')) and row.get('Videographer'):
                        response += f"• Videographer: {row.get('Videographer')}\n"
                    if pd.notna(row.get('Current Version')) and row.get('Current Version'):
                        response += f"• Current Version: {row.get('Current Version')}\n"
                else:
                    # Detailed format - all fields
                    response += f"**Task #{task_num}** - {status}\n"
                    for col in df.columns:
                        val = row.get(col)
                        if pd.notna(val) and str(val).strip():
                            response += f"• {col}: {val}\n"
                
                response += "\n"
        else:
            response += "_No live tasks currently in Excel_\n\n"
        
        # 2. COMPLETED TASKS FROM HISTORY DATABASE
        if include_history:
            response += "-" * 50 + "\n\n"
            response += "**✅ COMPLETED TASKS (History DB)**\n\n"
            
            try:
                with sqlite3.connect(HISTORY_DB_PATH) as conn:
                    # Get count
                    cursor = conn.execute("SELECT COUNT(*) FROM completed_tasks")
                    history_count = cursor.fetchone()[0]
                    response += f"_Total completed tasks: {history_count}_\n\n"
                    
                    if history_count > 0:
                        # Get recent completed tasks
                        if format == "summary":
                            query = """
                                SELECT task_number, brand, campaign_start_date, campaign_end_date,
                                       reference_number, location, sales_person, videographer,
                                       status, completed_at
                                FROM completed_tasks
                                ORDER BY task_number DESC
                                LIMIT 20
                            """
                        else:
                            query = """
                                SELECT *
                                FROM completed_tasks
                                ORDER BY task_number DESC
                                LIMIT 20
                            """
                        
                        cursor = conn.execute(query)
                        columns = [desc[0] for desc in cursor.description]
                        
                        rows = cursor.fetchall()
                        response += f"_Showing most recent {len(rows)} completed tasks:_\n\n"
                        
                        for row in rows:
                            task_dict = dict(zip(columns, row))
                            task_num = task_dict.get('task_number', 'N/A')
                            
                            if format == "summary":
                                response += f"**Task #{task_num}** - Completed\n"
                                response += f"• Brand: {task_dict.get('brand', 'N/A')}\n"
                                response += f"• Campaign: {task_dict.get('campaign_start_date', 'N/A')} to {task_dict.get('campaign_end_date', 'N/A')}\n"
                                response += f"• Reference: {task_dict.get('reference_number', 'N/A')}\n"
                                response += f"• Location: {task_dict.get('location', 'N/A')}\n"
                                response += f"• Sales Person: {task_dict.get('sales_person', 'N/A')}\n"
                                response += f"• Videographer: {task_dict.get('videographer', 'N/A')}\n"
                                response += f"• Completed: {task_dict.get('completed_at', 'N/A')}\n"
                            else:
                                response += f"**Task #{task_num}** - Completed\n"
                                for col, val in task_dict.items():
                                    if val and str(val).strip() and col != 'id':
                                        response += f"• {col}: {val}\n"
                            
                            response += "\n"
                    else:
                        response += "_No completed tasks in history database_\n\n"
                        
            except Exception as e:
                logger.error(f"Error reading history database: {e}")
                response += f"_Error reading history database: {str(e)}_\n\n"
        
        # 3. SUMMARY STATISTICS
        response += "=" * 50 + "\n"
        response += "**📈 SUMMARY STATISTICS**\n\n"
        
        # Status breakdown from Excel
        if len(df) > 0:
            status_counts = df['Status'].value_counts()
            response += "**Live Task Status Breakdown:**\n"
            for status, count in status_counts.items():
                response += f"• {status}: {count}\n"
            response += "\n"
            
            # Videographer workload
            videographer_counts = df[df['Videographer'].notna()]['Videographer'].value_counts()
            if len(videographer_counts) > 0:
                response += "**Videographer Assignments:**\n"
                for videographer, count in videographer_counts.items():
                    if videographer:
                        response += f"• {videographer}: {count} tasks\n"
                response += "\n"
        
        # History stats
        if include_history:
            try:
                with sqlite3.connect(HISTORY_DB_PATH) as conn:
                    cursor = conn.execute("SELECT COUNT(*) FROM completed_tasks")
                    total_completed = cursor.fetchone()[0]
                    
                    # Get completion by month
                    cursor = conn.execute("""
                        SELECT 
                            substr(completed_at, 4, 7) as month,
                            COUNT(*) as count
                        FROM completed_tasks
                        WHERE completed_at IS NOT NULL
                        GROUP BY substr(completed_at, 4, 7)
                        ORDER BY substr(completed_at, 7, 4) DESC, substr(completed_at, 4, 2) DESC
                        LIMIT 6
                    """)
                    
                    monthly_stats = cursor.fetchall()
                    if monthly_stats:
                        response += "**Recent Monthly Completions:**\n"
                        for month, count in monthly_stats:
                            response += f"• {month}: {count} completed\n"
                        response += "\n"
                        
            except Exception as e:
                logger.warning(f"Error getting history stats: {e}")
        
        response += f"\n_Export generated at {datetime.now(UAE_TZ).strftime('%d-%m-%Y %H:%M:%S')} UAE Time_"
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting data: {e}")
        return f"❌ Error exporting data: {str(e)}"

async def get_task_by_number(task_number: int) -> Dict[str, Any]:
    """Get a specific task by task number"""
    try:
        df = await read_excel_async()
        
        # Find the task
        task_row = df[df['Task #'] == task_number]
        
        if task_row.empty:
            return None
        
        # Convert to dictionary
        task_data = task_row.iloc[0].to_dict()
        
        # Convert timestamps to string for JSON serialization
        for key, value in task_data.items():
            if pd.isna(value):
                task_data[key] = ""
            elif hasattr(value, 'strftime'):
                task_data[key] = value.strftime('%d-%m-%Y')
                
        return task_data
    except Exception as e:
        logger.error(f"Error getting task {task_number}: {e}")
        return None

async def get_next_task_number() -> int:
    """Get the next available task number from both Excel and historical DB"""
    logger.info("\n=== GENERATING NEW TASK NUMBER ===")
    max_excel_task = 0
    max_db_task = 0
    excel_tasks = []
    
    try:
        # Get max from Excel
        df = await read_excel_async()
        logger.info(f"Excel file has {len(df)} rows")
        if len(df) > 0:
            # Get all task numbers for debugging
            excel_tasks = df['Task #'].dropna().tolist()
            logger.info(f"Task numbers in Excel: {sorted([int(t) for t in excel_tasks if str(t).isdigit()])[:10]}...")
            excel_max = df['Task #'].max()
            if pd.notna(excel_max):
                max_excel_task = int(excel_max)
            logger.info(f"Maximum task number in Excel: {max_excel_task}")
        else:
            logger.info("Excel file is empty")
    except Exception as e:
        logger.warning(f"Error reading Excel for task number: {e}")
    
    try:
        # Get max from SQLite history database
        with sqlite3.connect(HISTORY_DB_PATH) as conn:
            # Get count and some samples
            cursor = conn.execute("SELECT COUNT(*) FROM completed_tasks")
            count = cursor.fetchone()[0]
            logger.info(f"History database has {count} completed tasks")
            
            # Get sample of task numbers
            cursor = conn.execute("""
                SELECT task_number FROM completed_tasks 
                WHERE task_number IS NOT NULL 
                ORDER BY task_number DESC LIMIT 10
            """)
            db_sample = [row[0] for row in cursor.fetchall()]
            if db_sample:
                logger.info(f"Recent task numbers in history DB: {db_sample}")
            
            # Get max
            cursor = conn.execute("""
                SELECT MAX(task_number) FROM completed_tasks
            """)
            db_max = cursor.fetchone()[0]
            if db_max is not None:
                max_db_task = int(db_max)
            logger.info(f"Maximum task number in history DB: {max_db_task}")
    except Exception as e:
        logger.warning(f"Error reading history for task number: {e}")
    
    # Return the highest task number + 1
    max_task = max(max_excel_task, max_db_task)
    next_task = max_task + 1 if max_task > 0 else 1
    logger.info(f"Next task number will be: {next_task} (Excel max: {max_excel_task}, DB max: {max_db_task})")
    logger.info("=== END TASK NUMBER GENERATION ===\n")
    return next_task

async def check_duplicate_reference(reference_number: str) -> Dict[str, Any]:
    """Check if reference number already exists in Excel or historical DB"""
    logger.info(f"\n=== CHECKING DUPLICATE REFERENCE: {reference_number} ===")
    try:
        # Clean reference number for comparison
        clean_ref = reference_number.replace("_", "-")
        logger.info(f"Cleaned reference: {clean_ref}")
        
        # Check in active Excel first
        df = await read_excel_async()
        logger.info(f"Checking {len(df)} rows in Excel")
        existing = df[df['Reference Number'] == clean_ref]
        
        if len(existing) > 0:
            logger.info(f"Found duplicate in Excel: Task #{existing.iloc[0]['Task #']}")
            # Get details of existing entry
            existing_entry = existing.iloc[0]
            
            # Safely get values with fallbacks
            result = {
                "is_duplicate": True,
                "existing_entry": {
                    "task_number": str(existing_entry.get('Task #', '')),
                    "brand": str(existing_entry.get('Brand', '')),
                    "start_date": str(existing_entry.get('Campaign Start Date', '')),
                    "end_date": str(existing_entry.get('Campaign End Date', '')),
                    "location": str(existing_entry.get('Location', '')),
                    "submitted_by": str(existing_entry.get('Submitted By', '')),
                    "timestamp": str(existing_entry.get('Timestamp', '')),
                    "status": "Active"
                }
            }
            
            # Format dates to DD-MM-YYYY
            date_fields = ['start_date', 'end_date']
            for field in date_fields:
                value = existing_entry.get(field.replace('_', ' ').title().replace('Start', 'Campaign Start').replace('End', 'Campaign End'), '')
                if pd.notna(value) and hasattr(value, 'strftime'):
                    result["existing_entry"][field] = value.strftime('%d-%m-%Y')
            
            # For backward compatibility, also include 'date' field
            result["existing_entry"]["date"] = result["existing_entry"]["start_date"]
            
            return result
        
        # Check in historical database
        logger.info("No duplicate found in Excel, checking history database...")
        try:
            with sqlite3.connect(HISTORY_DB_PATH) as conn:
                # First count how many completed tasks we have
                cursor = conn.execute("SELECT COUNT(*) FROM completed_tasks")
                total_history = cursor.fetchone()[0]
                logger.info(f"Checking {total_history} completed tasks in history")
                
                cursor = conn.execute("""
                    SELECT task_number, brand, campaign_start_date, campaign_end_date, 
                           location, submitted_by, completed_at
                    FROM completed_tasks 
                    WHERE reference_number = ?
                """, (clean_ref,))
                
                row = cursor.fetchone()
                if row:
                    logger.info(f"Found duplicate in history: Task #{row[0]} (Completed: {row[6]})")
                    return {
                        "is_duplicate": True,
                        "existing_entry": {
                            "task_number": str(row[0]),
                            "brand": str(row[1] or ''),
                            "start_date": str(row[2] or ''),
                            "end_date": str(row[3] or ''),
                            "location": str(row[4] or ''),
                            "submitted_by": str(row[5] or ''),
                            "timestamp": str(row[6] or ''),
                            "date": str(row[2] or ''),  # For backward compatibility
                            "status": "Archived (Completed)"
                        }
                    }
        except Exception as e:
            logger.warning(f"Error checking historical data for duplicate: {e}")
        
        logger.info(f"No duplicate found for reference: {clean_ref}")
        logger.info("=== END DUPLICATE CHECK ===\n")
        return {"is_duplicate": False}
        
    except Exception as e:
        logger.error(f"Error checking duplicate reference: {e}")
        return {"is_duplicate": False}


def archive_completed_task(row: Dict[str, Any]) -> None:
    try:
        # Use SQLite for history
        with sqlite3.connect(HISTORY_DB_PATH) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS completed_tasks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_number INTEGER,
                    brand TEXT,
                    campaign_start_date TEXT,
                    campaign_end_date TEXT,
                    reference_number TEXT,
                    location TEXT,
                    sales_person TEXT,
                    submitted_by TEXT,
                    status TEXT,
                    filming_date TEXT,
                    videographer TEXT,
                    current_version TEXT,
                    version_history TEXT,
                    pending_timestamps TEXT,
                    submitted_timestamps TEXT,
                    returned_timestamps TEXT,
                    rejected_timestamps TEXT,
                    accepted_timestamps TEXT,
                    completed_at TEXT
                );
            """)
            # Safe getters
            def g(key: str) -> str:
                val = row.get(key, '')
                try:
                    return '' if val is None else str(val)
                except Exception:
                    return ''
            task_number_val = row.get('Task #', 0)
            task_number_int = int(task_number_val) if str(task_number_val).isdigit() else None
            conn.execute("""
                INSERT INTO completed_tasks
                (task_number, brand, campaign_start_date, campaign_end_date, reference_number, location, sales_person, submitted_by, status, filming_date, videographer, current_version, version_history, pending_timestamps, submitted_timestamps, returned_timestamps, rejected_timestamps, accepted_timestamps, completed_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                task_number_int,
                g('Brand'),
                g('Campaign Start Date'),
                g('Campaign End Date'),
                g('Reference Number'),
                g('Location'),
                g('Sales Person'),
                g('Submitted By'),
                g('Status'),
                g('Filming Date'),
                g('Videographer'),
                g('Current Version'),
                g('Version History'),
                g('Pending Timestamps'),
                g('Submitted Timestamps'),
                g('Returned Timestamps'),
                g('Rejected Timestamps'),
                g('Accepted Timestamps'),
                datetime.now(UAE_TZ).strftime('%d-%m-%Y %H:%M:%S')
            ))
    except Exception as e:
        logger.error(f"Failed to archive task: {e}")

async def log_design_request(brand: str, campaign_start_date: str, campaign_end_date: str, reference_number: str, 
                              location: str = "", sales_person: str = "", submitted_by: str = ""):
    """Log a design request to Excel"""
    data = {
        "brand": brand,
        "start_date": campaign_start_date,
        "end_date": campaign_end_date,
        "reference_number": reference_number,
        "location": location,
        "sales_person": sales_person,
        "submitted_by": submitted_by
    }
    
    result = await save_to_excel(data)
    if result["success"]:
        task_number = result["task_number"]
        return f"✅ **Task #{task_number} created successfully!**\n\n• Brand: {brand}\n• Start Date: {campaign_start_date}\n• End Date: {campaign_end_date}\n• Reference: {reference_number}\n• Location: {location}\n• Sales Person: {sales_person}"
    else:
        return "I'm sorry, but there was an error saving the request. Could you please try again?"


async def update_movement_timestamp(task_number: int, folder: str, version: Optional[int] = None):
    """Update movement timestamp for a task when video moves to a new folder; if version is provided, tag it."""
    try:
        df = await read_excel_async()
        
        # Find the task
        task_index = df[df['Task #'] == task_number].index
        if task_index.empty:
            logger.error(f"Task #{task_number} not found for timestamp update")
            return
        
        idx = task_index[0]
        
        # Map folder names to timestamp columns (using actual status names)
        folder_to_column = {
            "pending": "Pending Timestamps",
            "Pending": "Pending Timestamps",  # Add capital P version
            "Critique": "Pending Timestamps",
            "submitted": "Submitted Timestamps", 
            "Submitted to Sales": "Submitted Timestamps",
            "returned": "Returned Timestamps",
            "Returned": "Returned Timestamps",
            "rejected": "Rejected Timestamps",
            "Rejected": "Rejected Timestamps",  # Add capital R version
            "Editing": "Rejected Timestamps",
            "accepted": "Accepted Timestamps",
            "Accepted": "Accepted Timestamps",  # Add capital A version
            "Done": "Accepted Timestamps"
        }
        
        column_name = folder_to_column.get(folder)
        if not column_name:
            logger.warning(f"Unknown folder for timestamp tracking: {folder}")
            return
        
        # Ensure column exists
        if column_name not in df.columns:
            df[column_name] = ""
        
        # Get current timestamps (stored as string)
        current_timestamps = str(df.at[idx, column_name]) if pd.notna(df.at[idx, column_name]) else ""
        
        # Parse existing timestamps
        if current_timestamps and current_timestamps != "":
            timestamps_list = current_timestamps.split("; ")
        else:
            timestamps_list = []
        
        # Add new timestamp (optionally version-tagged)
        new_timestamp = datetime.now(UAE_TZ).strftime("%d-%m-%Y %H:%M:%S")
        entry = f"v{version}:{new_timestamp}" if version is not None else new_timestamp
        timestamps_list.append(entry)
        
        # Join back to string
        df.at[idx, column_name] = "; ".join(timestamps_list)
        
        # Optionally set current version
        if version is not None:
            if 'Current Version' not in df.columns:
                df['Current Version'] = ""
            df.at[idx, 'Current Version'] = version
        
        # Use safe write to avoid concurrency issues
        from excel_lock_utils import safe_write_excel
        success = await safe_write_excel(df)
        
        if not success:
            logger.error(f"Failed to update movement timestamp - file may be locked")
            return
        
        logger.info(f"✅ Updated movement timestamp for Task #{task_number} in {folder} (v{version if version is not None else '-'})")
    except Exception as e:
        logger.error(f"Error updating movement timestamp: {e}")

async def delete_task_by_number(task_number: int) -> Dict[str, Any]:
    """Delete a task by task number"""
    try:
        df = await read_excel_async()
        
        # Find the task
        task_index = df[df['Task #'] == task_number].index
        
        if task_index.empty:
            return {"success": False, "error": "Task not found"}
        
        # Get task data before deleting
        task_data = df.iloc[task_index[0]].to_dict()
        
        # Archive the task before deleting
        archive_completed_task(task_data)
        
        # Remove from DataFrame
        df = df.drop(task_index)
        
        # Use safe write to avoid concurrency issues
        from excel_lock_utils import safe_write_excel
        success = await safe_write_excel(df)
        
        if not success:
            logger.error(f"Failed to delete task - file may be locked")
            return {"success": False, "error": "Excel file is locked"}
        
        # If task was assigned, archive/remove the Trello card
        if str(task_data.get('Status', '')).startswith('Assigned to'):
            from trello_utils import archive_trello_card, get_trello_card_by_task_number
            card = await get_trello_card_by_task_number(task_number)
            if card:
                archived = archive_trello_card(card['id'])
                if not archived:
                    logger.warning(f"Could not archive Trello card for task {task_number}")
            else:
                logger.warning(f"Could not find Trello card for task {task_number}")
        
        return {"success": True, "task_data": task_data}
        
    except Exception as e:
        logger.error(f"Error deleting task {task_number}: {e}")
        return {"success": False, "error": str(e)}

async def update_task_by_number(task_number: int, updates: Dict[str, Any], current_data: Dict[str, Any] = None) -> Dict[str, Any]:
    """Update a task by task number, including Trello if already assigned"""
    try:
        df = await read_excel_async()
        
        # Find the task index
        task_index = df[df['Task #'] == task_number].index
        
        if task_index.empty:
            return {"success": False, "error": "Task not found"}
        
        idx = task_index[0]
        
        # Get current data if not provided
        if not current_data:
            current_data = df.iloc[idx].to_dict()
            
        # Format dates for display (convert from pandas datetime to DD-MM-YYYY string)
        date_fields = ['Campaign Start Date', 'Campaign End Date', 'Filming Date', 'Timestamp']
        for field in date_fields:
            if field in current_data and pd.notna(current_data[field]):
                # Convert to datetime if it's not already
                if isinstance(current_data[field], str):
                    try:
                        date_obj = pd.to_datetime(current_data[field])
                        current_data[field] = date_obj.strftime('%d-%m-%Y')
                    except:
                        pass  # Keep original if conversion fails
                else:
                    # It's already a datetime object
                    current_data[field] = current_data[field].strftime('%d-%m-%Y')
        
        # Check if task is already assigned (has Trello card)
        is_assigned = str(current_data.get('Status', '')).startswith('Assigned to')
        trello_updates_needed = False
        trello_updates = {}
        
        # If assigned, prepare Trello updates
        if is_assigned:
            # Check if videographer changed - this means reassignment
            if 'Videographer' in updates and updates['Videographer'] != current_data.get('Videographer'):
                # Update Status to reflect new assignee
                updates['Status'] = f"Assigned to {updates['Videographer']}"
                trello_updates_needed = True
                trello_updates['assignee'] = updates['Videographer']
            
            # Check what needs updating in Trello
            if 'Status' in updates and updates['Status'] != current_data.get('Status'):
                # Status change means reassignment
                trello_updates_needed = True
                new_assignee = updates['Status'].replace('Assigned to ', '')
                trello_updates['assignee'] = new_assignee
            
            if 'Filming Date' in updates and updates['Filming Date'] != current_data.get('Filming Date'):
                # Date change - need to update checklist dates
                trello_updates_needed = True
                # Store new filming date for checklist update
                try:
                    filming_date = pd.to_datetime(updates['Filming Date'])
                    trello_updates['filming_date'] = filming_date
                except:
                    pass
            
            # Check if details changed that require description update
            detail_fields = ['Brand', 'Campaign Start Date', 'Campaign End Date', 'Reference Number', 'Location', 'Sales Person']
            details_changed = any(field in updates and updates[field] != current_data.get(field) for field in detail_fields)
            
            if details_changed or 'Videographer' in updates:
                trello_updates_needed = True
                # We'll update description after updating Excel
        
        # Update Excel fields
        for field, value in updates.items():
            if field in df.columns and field != 'Task #':  # Don't allow changing task number
                df.at[idx, field] = value
        
        # Update timestamp
        df.at[idx, 'Timestamp'] = datetime.now(UAE_TZ).strftime('%d-%m-%Y %H:%M:%S')
        
        # Use safe write to avoid concurrency issues
        from excel_lock_utils import safe_write_excel
        success = await safe_write_excel(df)
        
        if not success:
            logger.error(f"Failed to update task - file may be locked")
            return {"success": False, "error": "Excel file is locked"}
        
        # If any task just moved to Done/Accepted, archive it, then remove from live sheet (optional)
        try:
            if updates.get('Status') == 'Done' or updates.get('Status') == 'Accepted':
                # Archive the completed row
                archive_completed_task(df.iloc[idx].to_dict())
                # Optionally, remove from live Excel (keep if you prefer)
                # df.drop(index=idx, inplace=True)
                # df.to_excel(EXCEL_FILE_PATH, index=False)
                # wb = load_workbook(EXCEL_FILE_PATH); ws = wb.active
                # for cell in ws[1]: cell.font = Font(bold=True); wb.save(EXCEL_FILE_PATH)
        except Exception as e:
            logger.warning(f"Archival step skipped: {e}")
        
        # Now handle Trello updates if needed
        if is_assigned and trello_updates_needed:
            # Get the Trello card
            trello_card = await get_trello_card_by_task_number(task_number)
            
            if trello_card:
                # Build card description with updated data
                updated_data = current_data.copy()
                updated_data.update(updates)
                
                description = f"""Task #{task_number}
Brand: {updated_data.get('Brand', '')}
Campaign Start Date: {updated_data.get('Campaign Start Date', '')}
Campaign End Date: {updated_data.get('Campaign End Date', '')}
Reference: {updated_data.get('Reference Number', '')}
Location: {updated_data.get('Location', '')}
Sales Person: {updated_data.get('Sales Person', '')}
Videographer: {updated_data.get('Videographer', '')}
Filming Date: {updated_data.get('Filming Date', '')}"""
                
                trello_updates['desc'] = description
                
                # Update card title if brand or reference changed
                if 'Brand' in updates or 'Reference Number' in updates:
                    new_brand = updated_data.get('Brand', '')
                    new_reference = updated_data.get('Reference Number', '')
                    new_title = f"Task #{task_number}: {new_brand} - {new_reference}"
                    trello_updates['name'] = new_title
                
                # If status changed (reassignment), we need to move the card
                if 'assignee' in trello_updates:
                    # Get list mapping
                    lists = await get_trello_lists()
                    new_list_id = lists.get(trello_updates['assignee'])
                    if new_list_id:
                        trello_updates['idList'] = new_list_id
                    del trello_updates['assignee']  # Remove as it's not a direct field
                
                # Handle filming date update separately if needed
                if 'filming_date' in trello_updates:
                    filming_date = trello_updates.pop('filming_date')
                    # Update checklist dates
                    from trello_utils import update_checklist_dates
                    checklist_updated = await update_checklist_dates(trello_card['id'], filming_date)
                    if not checklist_updated:
                        logger.warning(f"Failed to update checklist dates for Task #{task_number}")
                
                # Update the Trello card with remaining updates
                if trello_updates:  # Only update if there are remaining updates
                    success = await update_trello_card(trello_card['id'], trello_updates)
                    if not success:
                        return {"success": True, "warning": "Excel updated but Trello update failed"}
                else:
                    success = True
            else:
                return {"success": True, "warning": "Excel updated but Trello card not found"}
        
        return {"success": True, "updates": updates}
        
    except Exception as e:
        logger.error(f"Error updating task {task_number}: {e}")
        return {"success": False, "error": str(e)}
